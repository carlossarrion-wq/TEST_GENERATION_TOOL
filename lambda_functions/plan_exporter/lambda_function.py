import json
import logging
import boto3
import uuid
from datetime import datetime, timedelta
from typing import Dict, Any, List
import os
import io
import base64
from decimal import Decimal

# Para generar archivos Excel
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None

# Para generar archivos PDF
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import inch
except ImportError:
    SimpleDocTemplate = None

logger = logging.getLogger()
logger.setLevel(logging.INFO)

dynamodb = boto3.resource('dynamodb')
s3_client = boto3.client('s3')

sessions_table = dynamodb.Table(os.environ.get('SESSIONS_TABLE', 'test-plan-sessions'))
s3_bucket = os.environ.get('S3_BUCKET', 'test-plan-exports')

def decimal_default(obj):
    """Helper function to convert Decimal objects to float for JSON serialization"""
    if isinstance(obj, Decimal):
        return float(obj)
    raise TypeError(f"Object of type {type(obj)} is not JSON serializable")

def lambda_handler(event: Dict[str, Any], context) -> Dict[str, Any]:
    """POST /export-plan - Exportar plan de pruebas en diferentes formatos"""
    logger.info("=== PLAN_EXPORTER STARTED ===")
    logger.info(f"Raw event received: {json.dumps(event, default=str)}")
    
    try:
        if event.get('httpMethod') == 'OPTIONS':
            logger.info("OPTIONS request detected, returning CORS response")
            return cors_response()
        
        logger.info("Processing POST request")
        
        # Manejo robusto del body - soporta API Gateway y invocación directa
        try:
            if 'body' in event:
                # Formato API Gateway
                if event['body'] is None:
                    return error_response(400, 'Request body is null')
                
                if isinstance(event['body'], str):
                    body = json.loads(event['body'])
                else:
                    body = event['body']
            else:
                # Invocación directa - el evento ES el body
                body = event
                logger.info("Direct invocation detected, using event as body")
                
        except json.JSONDecodeError as e:
            return error_response(400, f'Invalid JSON in request body: {str(e)}')
        except Exception as e:
            return error_response(400, f'Error parsing request body: {str(e)}')
        
        required_fields = ['session_id', 'format']
        missing_fields = [field for field in required_fields if field not in body]
        
        if missing_fields:
            return error_response(400, f'Missing required fields: {", ".join(missing_fields)}')
        
        session_id = body['session_id']
        export_format = body['format'].lower()
        include_metadata = body.get('include_metadata', True)
        
        # Validar formato
        valid_formats = ['excel', 'pdf', 'json', 'csv']
        if export_format not in valid_formats:
            return error_response(400, f'Invalid format. Must be one of: {", ".join(valid_formats)}')
        
        logger.info(f"Exporting plan for session {session_id} in format {export_format}")
        
        # Obtener datos de la sesión
        try:
            session_response = sessions_table.get_item(Key={'id': session_id})
            
            if 'Item' not in session_response:
                return error_response(404, f'Session not found: {session_id}')
            
            session_data = session_response['Item']
            logger.info(f"Session found: {session_data.get('id', 'unknown')}")
        except Exception as e:
            logger.error(f"Error retrieving session: {str(e)}")
            return error_response(500, 'Error retrieving session data')
        
        # Extraer plan de pruebas de la última iteración
        iterations = session_data.get('iterations', [])
        if not iterations:
            return error_response(400, 'No test plan found in session')
        
        # Obtener la última iteración con plan generado
        latest_iteration = None
        for iteration in reversed(iterations):
            if iteration.get('generated_plan') and iteration['generated_plan'].get('test_cases'):
                latest_iteration = iteration
                break
        
        if not latest_iteration:
            return error_response(400, 'No test cases found in session')
        
        test_plan = latest_iteration['generated_plan']
        plan_config = session_data.get('plan_configuration', {})
        
        logger.info(f"Found test plan with {len(test_plan.get('test_cases', []))} test cases")
        
        # Generar archivo según el formato
        file_content = None
        file_name = f"plan_{session_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        content_type = 'application/octet-stream'
        
        if export_format == 'excel':
            file_content, content_type = generate_excel_file(test_plan, plan_config, include_metadata)
            file_name += '.xlsx'
        elif export_format == 'pdf':
            file_content, content_type = generate_pdf_file(test_plan, plan_config, include_metadata)
            file_name += '.pdf'
        elif export_format == 'json':
            file_content, content_type = generate_json_file(test_plan, plan_config, include_metadata)
            file_name += '.json'
        elif export_format == 'csv':
            file_content, content_type = generate_csv_file(test_plan, plan_config, include_metadata)
            file_name += '.csv'
        
        if file_content is None:
            return error_response(500, f'Failed to generate {export_format} file')
        
        # Subir archivo a S3
        try:
            s3_key = f"exports/{session_id}/{file_name}"
            
            s3_client.put_object(
                Bucket=s3_bucket,
                Key=s3_key,
                Body=file_content,
                ContentType=content_type,
                ContentDisposition=f'attachment; filename="{file_name}"'
            )
            
            # Generar URL de descarga con expiración
            download_url = s3_client.generate_presigned_url(
                'get_object',
                Params={'Bucket': s3_bucket, 'Key': s3_key},
                ExpiresIn=3600  # 1 hora
            )
            
            expires_at = (datetime.utcnow() + timedelta(hours=1)).isoformat()
            
            logger.info(f"File uploaded to S3: {s3_key}")
            
            return success_response({
                'download_url': download_url,
                'file_name': file_name,
                'expires_at': expires_at,
                'format': export_format,
                'file_size': len(file_content) if isinstance(file_content, (bytes, str)) else 0
            })
            
        except Exception as e:
            logger.error(f"Error uploading to S3: {str(e)}")
            # Fallback: devolver archivo como base64
            if isinstance(file_content, str):
                file_content = file_content.encode('utf-8')
            
            file_base64 = base64.b64encode(file_content).decode('utf-8')
            
            return success_response({
                'download_url': f'data:{content_type};base64,{file_base64}',
                'file_name': file_name,
                'expires_at': expires_at,
                'format': export_format,
                'file_size': len(file_content)
            })
            
    except Exception as e:
        logger.error(f"Export failed: {str(e)}")
        return error_response(500, 'Export failed', str(e))

def generate_excel_file(test_plan: Dict, plan_config: Dict, include_metadata: bool) -> tuple:
    """Generar archivo Excel con los casos de prueba"""
    if openpyxl is None:
        logger.error("openpyxl not available, cannot generate Excel file")
        return None, None
    
    try:
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Plan de Pruebas"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Información del plan
        if include_metadata:
            ws['A1'] = "PLAN DE PRUEBAS"
            ws['A1'].font = Font(bold=True, size=16)
            ws.merge_cells('A1:H1')
            
            ws['A3'] = "Título:"
            ws['B3'] = plan_config.get('plan_title', 'Sin título')
            ws['A4'] = "Tipo:"
            ws['B4'] = plan_config.get('plan_type', 'No especificado')
            ws['A5'] = "Cobertura objetivo:"
            ws['B5'] = f"{plan_config.get('coverage_percentage', 0)}%"
            ws['A6'] = "Fecha de generación:"
            ws['B6'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            start_row = 8
        else:
            start_row = 1
        
        # Headers de la tabla
        headers = [
            "Número", "Nombre del Caso", "Descripción", "Precondiciones",
            "Pasos de Prueba", "Resultados Esperados", "Prioridad", "Estado"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Datos de los casos de prueba
        test_cases = test_plan.get('test_cases', [])
        for row, test_case in enumerate(test_cases, start_row + 1):
            ws.cell(row=row, column=1, value=test_case.get('testcase_number', row - start_row))
            ws.cell(row=row, column=2, value=test_case.get('test_case_name', ''))
            ws.cell(row=row, column=3, value=test_case.get('test_case_description', ''))
            ws.cell(row=row, column=4, value=test_case.get('preconditions', ''))
            
            # Convertir pasos a texto
            steps = test_case.get('test_steps', [])
            if isinstance(steps, list):
                steps_text = '\n'.join([f"{i+1}. {step}" for i, step in enumerate(steps)])
            else:
                steps_text = str(steps)
            ws.cell(row=row, column=5, value=steps_text)
            
            ws.cell(row=row, column=6, value=test_case.get('expected_results', ''))
            ws.cell(row=row, column=7, value=test_case.get('priority', 'MEDIA'))
            ws.cell(row=row, column=8, value=test_case.get('status', 'PROPOSED'))
            
            # Aplicar bordes
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).border = border
        
        # Ajustar ancho de columnas
        column_widths = [10, 25, 35, 20, 30, 25, 12, 12]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        
    except Exception as e:
        logger.error(f"Error generating Excel file: {str(e)}")
        return None, None

def generate_pdf_file(test_plan: Dict, plan_config: Dict, include_metadata: bool) -> tuple:
    """Generar archivo PDF con los casos de prueba"""
    if SimpleDocTemplate is None:
        logger.error("reportlab not available, cannot generate PDF file")
        return None, None
    
    try:
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4)
        story = []
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # Center
        )
        
        # Título
        story.append(Paragraph("PLAN DE PRUEBAS", title_style))
        story.append(Spacer(1, 20))
        
        # Información del plan
        if include_metadata:
            info_data = [
                ['Título:', plan_config.get('plan_title', 'Sin título')],
                ['Tipo:', plan_config.get('plan_type', 'No especificado')],
                ['Cobertura objetivo:', f"{plan_config.get('coverage_percentage', 0)}%"],
                ['Fecha de generación:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
                ['Total de casos:', str(len(test_plan.get('test_cases', [])))]
            ]
            
            info_table = Table(info_data, colWidths=[2*inch, 4*inch])
            info_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(info_table)
            story.append(Spacer(1, 30))
        
        # Tabla de casos de prueba
        story.append(Paragraph("CASOS DE PRUEBA", styles['Heading2']))
        story.append(Spacer(1, 10))
        
        # Headers
        headers = ['#', 'Nombre', 'Descripción', 'Prioridad']
        table_data = [headers]
        
        # Datos
        test_cases = test_plan.get('test_cases', [])
        for i, test_case in enumerate(test_cases, 1):
            row = [
                str(i),
                test_case.get('test_case_name', '')[:30] + '...' if len(test_case.get('test_case_name', '')) > 30 else test_case.get('test_case_name', ''),
                test_case.get('test_case_description', '')[:50] + '...' if len(test_case.get('test_case_description', '')) > 50 else test_case.get('test_case_description', ''),
                test_case.get('priority', 'MEDIA')
            ]
            table_data.append(row)
        
        table = Table(table_data, colWidths=[0.5*inch, 2*inch, 3*inch, 1*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'TOP')
        ]))
        
        story.append(table)
        
        doc.build(story)
        output.seek(0)
        
        return output.getvalue(), 'application/pdf'
        
    except Exception as e:
        logger.error(f"Error generating PDF file: {str(e)}")
        return None, None

def generate_json_file(test_plan: Dict, plan_config: Dict, include_metadata: bool) -> tuple:
    """Generar archivo JSON con los casos de prueba"""
    try:
        export_data = {
            'test_plan': test_plan,
            'exported_at': datetime.utcnow().isoformat(),
            'format': 'json'
        }
        
        if include_metadata:
            export_data['plan_configuration'] = plan_config
        
        json_content = json.dumps(export_data, indent=2, default=decimal_default, ensure_ascii=False)
        
        return json_content.encode('utf-8'), 'application/json'
        
    except Exception as e:
        logger.error(f"Error generating JSON file: {str(e)}")
        return None, None

def generate_csv_file(test_plan: Dict, plan_config: Dict, include_metadata: bool) -> tuple:
    """Generar archivo CSV con los casos de prueba"""
    try:
        import csv
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Headers
        headers = [
            'Número', 'Nombre del Caso', 'Descripción', 'Precondiciones',
            'Pasos de Prueba', 'Resultados Esperados', 'Prioridad', 'Estado'
        ]
        writer.writerow(headers)
        
        # Datos
        test_cases = test_plan.get('test_cases', [])
        for i, test_case in enumerate(test_cases, 1):
            steps = test_case.get('test_steps', [])
            if isinstance(steps, list):
                steps_text = '; '.join(steps)
            else:
                steps_text = str(steps)
            
            row = [
                test_case.get('testcase_number', i),
                test_case.get('test_case_name', ''),
                test_case.get('test_case_description', ''),
                test_case.get('preconditions', ''),
                steps_text,
                test_case.get('expected_results', ''),
                test_case.get('priority', 'MEDIA'),
                test_case.get('status', 'PROPOSED')
            ]
            writer.writerow(row)
        
        csv_content = output.getvalue()
        output.close()
        
        return csv_content.encode('utf-8'), 'text/csv'
        
    except Exception as e:
        logger.error(f"Error generating CSV file: {str(e)}")
        return None, None

def success_response(data):
    return {
        'statusCode': 200,
        'headers': cors_headers(),
        'body': json.dumps({**data, 'timestamp': datetime.utcnow().isoformat()}, default=decimal_default)
    }

def error_response(status_code, message, details=None):
    return {
        'statusCode': status_code,
        'headers': cors_headers(),
        'body': json.dumps({
            'error': message,
            'details': details,
            'timestamp': datetime.utcnow().isoformat()
        }, default=decimal_default)
    }

def cors_response():
    return {'statusCode': 200, 'headers': cors_headers(), 'body': ''}

def cors_headers():
    return {
        'Content-Type': 'application/json',
        'Access-Control-Allow-Origin': '*',
        'Access-Control-Allow-Headers': 'Content-Type,X-Amz-Date,Authorization,X-Api-Key,X-Amz-Security-Token',
        'Access-Control-Allow-Methods': 'POST,OPTIONS'
    }
